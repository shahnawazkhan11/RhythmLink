"""
Import data from Excel files (.xls and .xlsx format)
"""
import openpyxl
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from django.utils import timezone
from django.db import transaction
from datetime import datetime, date

from artists.models import Genre, Artist, Album, Track
from events.models import Event, Venue, EventType, Performs
from customers.models import Customer, Booking, Ticket, FanInteraction


class Command(BaseCommand):
    help = 'Import data from Excel files (.xls and .xlsx format)'

    def add_arguments(self, parser):
        parser.add_argument('--folder', type=str, default='data', help='Folder containing Excel files')
        parser.add_argument('--clear', action='store_true', help='Clear existing data before import')

    def handle(self, *args, **options):
        folder = options['folder']
        clear_data = options['clear']
        
        if clear_data:
            self.stdout.write(self.style.WARNING('[!] Clearing existing data...'))
            self.clear_all_data()
        
        self.stdout.write(self.style.SUCCESS('[*] Starting Excel data import...'))
        
        try:
            # Import in dependency order
            self.import_artists(f'{folder}/Artists.xls')
            self.import_albums(f'{folder}/Albums.xls')
            self.import_tracks(f'{folder}/Tracks.xls')
            self.import_fans(f'{folder}/Fans.xls')
            self.import_events(f'{folder}/Events.xls')
            self.import_fan_interactions(f'{folder}/Fan_interactions.xls')
            
            # Check for ticket sales file
            try:
                self.import_ticket_sales(f'{folder}/Ticket_sales.xls')
            except FileNotFoundError:
                self.stdout.write(self.style.WARNING('  [!] Ticket_sales.xls not found, skipping...'))
            
            self.stdout.write(self.style.SUCCESS('\n[SUCCESS] Excel import complete!'))
            self.print_summary()
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'\n[ERROR] Import failed: {str(e)}'))
            raise

    def clear_all_data(self):
        """Clear all existing data"""
        FanInteraction.objects.all().delete()
        Booking.objects.all().delete()
        Ticket.objects.all().delete()
        Track.objects.all().delete()
        Album.objects.all().delete()
        Performs.objects.all().delete()
        Event.objects.all().delete()
        Customer.objects.all().delete()
        Artist.objects.all().delete()
        Genre.objects.all().delete()
        User.objects.filter(is_superuser=False).delete()
        self.stdout.write('  [+] Data cleared')

    def parse_date(self, value):
        """Parse date from Excel"""
        if isinstance(value, (date, datetime)):
            if isinstance(value, datetime):
                return value.date()
            return value
        if isinstance(value, str):
            # Try parsing string date
            for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%d/%m/%Y']:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return timezone.now().date()

    def parse_datetime(self, value):
        """Parse datetime from Excel"""
        if isinstance(value, datetime):
            if timezone.is_aware(value):
                return value
            return timezone.make_aware(value)
        if isinstance(value, date):
            dt = datetime.combine(value, datetime.min.time())
            return timezone.make_aware(dt)
        if isinstance(value, str):
            # Try parsing string datetime
            for fmt in ['%Y-%m-%d %H:%M:%S', '%d-%m-%Y %H:%M:%S', '%m/%d/%Y %H:%M:%S']:
                try:
                    dt = datetime.strptime(value, fmt)
                    return timezone.make_aware(dt)
                except ValueError:
                    continue
        return timezone.now()

    @transaction.atomic
    def import_artists(self, file_path):
        """Import artists from Excel"""
        self.stdout.write('\n[*] Importing artists...')
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Get header row
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
            
            # Get or create genre
            genre_name = str(row_dict.get('genres', 'Unknown')).strip()
            genre, _ = Genre.objects.get_or_create(name=genre_name)
            
            # Split name into first and last name
            full_name = str(row_dict['name']).strip()
            name_parts = full_name.split(' ', 1)
            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ''
            
            # Create artist
            artist, created = Artist.objects.update_or_create(
                first_name=first_name,
                last_name=last_name,
                defaults={
                    'genre': genre,
                    'followers': int(row_dict.get('followers', 0)) if row_dict.get('followers') else 0,
                    'popularity': int(row_dict.get('popularity', 0)) if row_dict.get('popularity') else 0,
                }
            )
            
            if created:
                full_name_display = f"{artist.first_name} {artist.last_name}".strip() if artist.last_name else artist.first_name
                self.stdout.write(f'  [+] Created artist: {full_name_display}')
        
        self.stdout.write(self.style.SUCCESS(f'  [OK] Imported {Artist.objects.count()} artists'))

    @transaction.atomic
    def import_albums(self, file_path):
        """Import albums from Excel"""
        self.stdout.write('\n[*] Importing albums...')
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
            
            # Find artist by ID or name
            artist_id = row_dict.get('artist_id')
            try:
                # Try to find artist by original position (row number in artists sheet)
                artist = Artist.objects.all()[int(artist_id) - 1] if artist_id else None
            except (IndexError, ValueError):
                self.stdout.write(self.style.WARNING(f'  [!] Artist ID {artist_id} not found'))
                continue
            
            release_date = self.parse_date(row_dict.get('release_date'))
            if not release_date:
                release_date = timezone.now().date()
            
            album, created = Album.objects.update_or_create(
                artist=artist,
                album_name=str(row_dict['album_name']).strip(),
                defaults={
                    'release_date': release_date,
                    'total_tracks': int(row_dict.get('total_tracks', 0)) if row_dict.get('total_tracks') else 0,
                }
            )
            
            if created:
                self.stdout.write(f'  [+] Created album: {album.album_name}')
        
        self.stdout.write(self.style.SUCCESS(f'  [OK] Imported {Album.objects.count()} albums'))

    @transaction.atomic
    def import_tracks(self, file_path):
        """Import tracks from Excel"""
        self.stdout.write('\n🎵 Importing tracks...')
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row = {headers[i]: sheet.cell_value(row_idx, i) for i in range(len(headers))}
            
            album_id = row.get('album_id')
            try:
                album = Album.objects.all()[int(album_id) - 1] if album_id else None
            except (IndexError, ValueError):
                self.stdout.write(self.style.WARNING(f'  ⚠️  Album ID {album_id} not found'))
                continue
            
            track, created = Track.objects.update_or_create(
                album=album,
                track_number=int(row.get('track_number', 1)),
                defaults={
                    'track_name': str(row['track_name']).strip(),
                    'duration_ms': int(row.get('duration_ms', 0)) if row.get('duration_ms') else 0,
                }
            )
            
            if created:
                self.stdout.write(f'  ✓ Created track: {track.track_name}')
        
        self.stdout.write(self.style.SUCCESS(f'  ✅ Imported {Track.objects.count()} tracks'))

    @transaction.atomic
    def import_fans(self, file_path):
        """Import fans/customers from Excel"""
        self.stdout.write('\n👥 Importing fans/customers...')
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row = {headers[i]: sheet.cell_value(row_idx, i) for i in range(len(headers))}
            
            fan_id = int(row.get('Fan_id', row_idx))
            name = str(row.get('name', f'Fan {fan_id}')).strip()
            email = str(row.get('email', f'fan{fan_id}@example.com')).strip()
            country = str(row.get('country', '')).strip()
            
            # Split name
            name_parts = name.split()
            first_name = name_parts[0] if name_parts else 'Fan'
            last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else str(fan_id)
            
            # Create Django user
            username = f'fan{fan_id}'
            user, user_created = User.objects.get_or_create(
                username=username,
                defaults={
                    'email': email,
                    'first_name': first_name,
                    'last_name': last_name,
                }
            )
            
            if user_created:
                user.set_password('fan123')
                user.save()
            
            # Create customer
            customer, created = Customer.objects.get_or_create(
                user=user,
                defaults={
                    'country': country,
                }
            )
            
            if created:
                self.stdout.write(f'  ✓ Created fan: {name} ({username})')
        
        self.stdout.write(self.style.SUCCESS(f'  ✅ Imported {Customer.objects.count()} fans'))

    @transaction.atomic
    def import_events(self, file_path):
        """Import events from Excel"""
        self.stdout.write('\n🎪 Importing events...')
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        headers = [cell.value for cell in sheet[1]]
        
        # Create default venue and event type
        default_venue, _ = Venue.objects.get_or_create(
            name='Main Arena',
            defaults={'location': 'Mumbai, India', 'capacity': 5000}
        )
        concert_type, _ = EventType.objects.get_or_create(
            name='Concert',
            defaults={'description': 'Live concert performance'}
        )
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row = {headers[i]: sheet.cell_value(row_idx, i) for i in range(len(headers))}
            
            artist_id = row.get('artist_id')
            try:
                artist = Artist.objects.all()[int(artist_id) - 1] if artist_id else None
            except (IndexError, ValueError):
                self.stdout.write(self.style.WARNING(f'  ⚠️  Artist ID {artist_id} not found'))
                continue
            
            event_date = self.parse_date(row.get('date'))
            if not event_date:
                event_date = timezone.now().date()
            
            location = str(row.get('location', 'TBD')).strip()
            
            # Create or get venue from location
            venue, _ = Venue.objects.get_or_create(
                name=f'{location} Venue',
                defaults={'location': location, 'capacity': 5000}
            )
            
            artist_full_name = f"{artist.first_name} {artist.last_name}".strip() if artist.last_name else artist.first_name
            
            event, created = Event.objects.update_or_create(
                name=str(row['name']).strip(),
                date=event_date,
                defaults={
                    'description': f'Live performance by {artist_full_name}',
                    'venue': venue,
                    'event_type': concert_type,
                    'status': 'upcoming' if event_date > timezone.now().date() else 'completed',
                }
            )
            
            # Create performance relationship
            Performs.objects.get_or_create(
                artist=artist,
                event=event
            )
            
            if created:
                self.stdout.write(f'  ✓ Created event: {event.name}')
        
        self.stdout.write(self.style.SUCCESS(f'  ✅ Imported {Event.objects.count()} events'))

    @transaction.atomic
    def import_fan_interactions(self, file_path):
        """Import fan interactions from Excel"""
        self.stdout.write('\n💖 Importing fan interactions...')
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        headers = [cell.value for cell in sheet[1]]
        
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row = {headers[i]: sheet.cell_value(row_idx, i) for i in range(len(headers))}
            
            fan_id = int(row.get('Fan_id', 0))
            track_id = int(row.get('track_id', 0))
            
            try:
                fan = Customer.objects.all()[fan_id - 1] if fan_id > 0 else None
                track = Track.objects.all()[track_id - 1] if track_id > 0 else None
                
                if not fan or not track:
                    continue
                
                interaction_type = str(row.get('type of interaction', 'play')).lower().strip()
                # Map interaction types
                type_mapping = {
                    'play': 'play',
                    'like': 'like',
                    'share': 'share',
                    'download': 'download',
                    'playlist': 'playlist_add',
                }
                interaction_type = type_mapping.get(interaction_type, 'play')
                
                timestamp = self.parse_datetime(row.get('time_stamp'))
                
                interaction, created = FanInteraction.objects.get_or_create(
                    fan=fan,
                    track=track,
                    timestamp=timestamp,
                    interaction_type=interaction_type,
                )
                
                if created:
                    count += 1
                    
            except (IndexError, ValueError) as e:
                continue
        
        self.stdout.write(self.style.SUCCESS(f'  ✅ Imported {count} fan interactions'))

    @transaction.atomic
    def import_ticket_sales(self, file_path):
        """Import ticket sales from Excel"""
        self.stdout.write('\n🎫 Importing ticket sales...')
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        headers = [cell.value for cell in sheet[1]]
        
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row = {headers[i]: sheet.cell_value(row_idx, i) for i in range(len(headers))}
            
            fan_id = int(row.get('Fan_id', 0))
            event_id = int(row.get('event_id', 0))
            
            try:
                fan = Customer.objects.all()[fan_id - 1] if fan_id > 0 else None
                event = Event.objects.all()[event_id - 1] if event_id > 0 else None
                
                if not fan or not event:
                    continue
                
                price = float(row.get('prices', 0)) if row.get('prices') else 0
                purchase_date = self.parse_datetime(row.get('purchase_date'))
                
                # Create booking
                booking, created = Booking.objects.get_or_create(
                    customer=fan,
                    event=event,
                    booking_date=purchase_date,
                    defaults={
                        'total_amount': price,
                        'status': 'confirmed',
                    }
                )
                
                if created:
                    # Create ticket
                    ticket = Ticket.objects.create(
                        event=event,
                        seat_number=f'S{row_idx}',
                        base_price=price,
                        final_price=price,
                        status='booked',
                    )
                    booking.tickets.add(ticket)
                    count += 1
                    
            except (IndexError, ValueError) as e:
                continue
        
        self.stdout.write(self.style.SUCCESS(f'  ✅ Imported {count} ticket sales'))

    def print_summary(self):
        """Print import summary"""
        self.stdout.write(self.style.SUCCESS('\n📊 Import Summary:'))
        self.stdout.write(f'  • Genres: {Genre.objects.count()}')
        self.stdout.write(f'  • Artists: {Artist.objects.count()}')
        self.stdout.write(f'  • Albums: {Album.objects.count()}')
        self.stdout.write(f'  • Tracks: {Track.objects.count()}')
        self.stdout.write(f'  • Events: {Event.objects.count()}')
        self.stdout.write(f'  • Venues: {Venue.objects.count()}')
        self.stdout.write(f'  • Fans/Customers: {Customer.objects.count()}')
        self.stdout.write(f'  • Fan Interactions: {FanInteraction.objects.count()}')
        self.stdout.write(f'  • Bookings: {Booking.objects.count()}')
        self.stdout.write(f'  • Tickets: {Ticket.objects.count()}')
